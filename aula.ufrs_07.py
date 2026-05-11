# -*- coding: utf-8 -*-
"""
Created on Mon May 11 19:21:55 2026

@author: Soraia Costa
"""
from openpyxl import load_workbook
import matplotlib.pyplot as gr
import numpy as np

wb=load_workbook("XYZW3.xlsx".data_only =True)
p=wb['Planilha1']

lin=p.max_row

op=np.zeros(lin)
hig=np.zeros(lin)
low=np.zeros(lin)
fec=np.zeros(lin)

for i in range(1,lin):
    op[i]=p.cell(row=i,column=2).value
    hig[i]=p.cell(row=i,column=3).value
    low[i]=p.cell(row=1,column=4).value
    fec[i]=p.cell(row=i,column=5).value
    
dados=[]

for i in range(lin):
    [dados.append([op[i],hig[i],low[i],fec[i]])]

gr.boxplot(dados)