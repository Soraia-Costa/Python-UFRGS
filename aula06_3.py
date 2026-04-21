# -*- coding: utf-8 -*-
"""
Created on Sat Apr 18 22:28:28 2026

@author: Soraia Costa
"""

from openpyxl import load_workbook

import statistics as st
import matplotlib.pyplot as gr

wb=load_workbook('dados.xlsx',data_only= True)
p=wb['Planilha1']

x=p['A']
y=p['B']

linha=p.max_row
coluna=p.max_column

dados=[]

for i in range(1,linha+1):
    dados.append(p.cell(row=i,column=2).value)
    
print(dados)

media=st.mean(dados)
desvio =st.pstdev(dados)

gr.subplot(211)

gr.hist(dados,bins=7)
gr.xlabel('Classes')
gr.ylabel('Freq')


gr.subplot(212)

gr.boxplot(dados)